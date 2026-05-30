"""
Crawler dvcqg dựa trên danh sách mã thủ tục.

Có 2 nguồn lấy danh sách mã TTHC:
  A. ONLINE (mặc định) — lấy động qua API, không cần file local:
     1. POST rest.jsp {service: procedure_get_list_agency_by_type_service_v2}
        → danh sách cơ quan (ID, AGENCY_NAME, CODE)
     2. GET export_exel_list_tthc.jsp?impl_agency_id=<ID>
        → file Excel danh sách thủ tục của cơ quan đó
     3. Parse Excel (bytes) → list mã TTHC
  B. LOCAL — đọc file .xlsx trong settings.XLSX_DATA_DIR (fallback/offline)

Sau khi có danh sách mã, với mỗi mã TTHC:
  a. POST rest.jsp {service: procedure_advanced_search_service_v2, keyword:<mã>}
     → lấy idTTHC
  b. GET export_word_detail_tthc.jsp?maTTHC=<code>&idTTHC=<id> → bytes .docx
  c. parse_docx → dict structured (xem dvcqg_docx_parser)
"""
from __future__ import annotations

import asyncio
import html as ihtml
import json
import re
from io import BytesIO
from pathlib import Path
from typing import AsyncIterator, Iterable

import httpx
from loguru import logger
from openpyxl import load_workbook

from app.core.config import settings
from app.crawler.parsers.dvcqg_docx_parser import parse_docx


REST_URL = "https://thutuc.dichvucong.gov.vn/jsp/rest.jsp"
EXPORT_URL = "https://thutuc.dichvucong.gov.vn/jsp/tthc/export/export_word_detail_tthc.jsp"
EXPORT_LIST_URL = "https://thutuc.dichvucong.gov.vn/jsp/tthc/export/export_exel_list_tthc.jsp"
# Trang chi tiết server-rendered — chứa link tải biểu mẫu (download_file.jsp)
DETAIL_HTML_URL = "https://thutuc.dichvucong.gov.vn/p/home/dvc-chi-tiet-thu-tuc-hanh-chinh.html"
# Regex bắt cặp (download_url, form_filename) trong bảng thành phần hồ sơ
_FORM_LINK_RE = re.compile(
    r'<a href="(https://csdl\.dichvucong\.gov\.vn/web/jsp/download_file\.jsp\?ma=[0-9a-f]+)"[^>]*>([^<]+)</a>'
)

_DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0 Safari/537.36",
    "Referer": "https://thutuc.dichvucong.gov.vn/jsp/rest.jsp",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "X-Requested-With": "XMLHttpRequest",
}


# ── XLSX reading ──────────────────────────────────────────────────────────────

# Procedure code looks like "1.015028", "2.000123" — digits.dot.digits, ≥ 6 digit tail.
_CODE_RE = re.compile(r"^\d+\.\d{4,}$")


def _find_header_row(ws) -> int:
    """Find the row containing 'Mã TTHC'. Returns 1-based row index, or -1."""
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if not row:
            continue
        for cell in row:
            if cell and "Mã TTHC" in str(cell):
                return i
    return -1


def _read_codes_from_worksheet(ws) -> list[dict]:
    """Core: extract code dicts from an openpyxl worksheet."""
    header_row = _find_header_row(ws)
    if header_row < 0:
        return []

    headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    col_idx = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        col_idx[str(h).strip()] = i

    code_col = col_idx.get("Mã TTHC")
    if code_col is None:
        return []

    out: list[dict] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or code_col >= len(row):
            continue
        code_val = row[code_col]
        if code_val is None:
            continue
        code = str(code_val).strip()
        if not _CODE_RE.match(code):
            continue
        out.append({
            "code": code,
            "name_xlsx": _val(row, col_idx.get("Tên TTHC")),
            "domain_xlsx": _val(row, col_idx.get("Lĩnh vực")),
            "object_xlsx": _val(row, col_idx.get("Đối tượng")),
            "agency_xlsx": _val(row, col_idx.get("Cơ quan công khai")),
            "level_xlsx": _val(row, col_idx.get("Cấp thực hiện")),
            "status_xlsx": _val(row, col_idx.get("Tình trạng")),
            "decision_xlsx": _val(row, col_idx.get("QĐ Công bố")),
        })
    return out


def read_codes_from_xlsx(path: Path) -> list[dict]:
    """Read a local .xlsx file → list of code dicts."""
    # KHÔNG dùng read_only=True: các xlsx của DVCQG có merged cells ở header
    # làm read_only mode chỉ thấy 4 rows thay vì hàng trăm. Bỏ flag → đọc đúng.
    wb = load_workbook(path, data_only=True)
    rows = _read_codes_from_worksheet(wb.active)
    if not rows:
        logger.warning(f"XLSX | no codes found | file={path.name}")
    return rows


def read_codes_from_bytes(xlsx_bytes: bytes) -> list[dict]:
    """Read xlsx bytes (vd tải online) → list of code dicts."""
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    return _read_codes_from_worksheet(wb.active)


def _val(row, idx) -> str | None:
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def list_xlsx_files(data_dir: Path | str | None = None) -> list[Path]:
    """Return list of .xlsx files in data_dir (excludes ~$ lock files)."""
    d = Path(data_dir or settings.XLSX_DATA_DIR)
    if not d.exists():
        logger.warning(f"XLSX | data dir does not exist | path={d}")
        return []
    return sorted([
        p for p in d.glob("*.xlsx")
        if not p.name.startswith("~$")
    ])


def collect_all_codes(data_dir: Path | str | None = None) -> list[dict]:
    """[LOCAL] Read all xlsx files and aggregate codes (deduplicated by code)."""
    files = list_xlsx_files(data_dir)
    seen: set[str] = set()
    out: list[dict] = []
    for f in files:
        try:
            rows = read_codes_from_xlsx(f)
        except Exception as e:
            logger.warning(f"XLSX | failed to read | file={f.name} | {e}")
            continue
        for r in rows:
            if r["code"] in seen:
                continue
            seen.add(r["code"])
            r["source_xlsx"] = f.name
            out.append(r)
    logger.info(f"XLSX | collected {len(out)} unique codes from {len(files)} files")
    return out


# ── ONLINE: lấy danh sách cơ quan + Excel động qua API ─────────────────────────

async def fetch_agency_list(client: httpx.AsyncClient) -> list[dict]:
    """
    POST rest.jsp → danh sách cơ quan cấp Bộ/ngành.
    Trả về list {id, name, code}.
    """
    payload = {
        "service": "procedure_get_list_agency_by_type_service_v2",
        "provider": "dvcquocgia",
        "type": "ref",
        "loaicoquan": "0",
    }
    try:
        r = await client.post(
            REST_URL,
            data={"params": json.dumps(payload, ensure_ascii=False)},
            headers=_DEFAULT_HEADERS,
            timeout=settings.CRAWLER_TIMEOUT,
        )
        r.raise_for_status()
        body = r.json()
    except Exception as e:
        logger.warning(f"Crawler | fetch_agency_list failed | {e}")
        return []

    out = []
    for a in body or []:
        aid = str(a.get("ID", "")).strip()
        if not aid:
            continue
        out.append({
            "id": aid,
            "name": (a.get("AGENCY_NAME") or a.get("NAME") or "").strip(),
            "code": (a.get("CODE") or "").strip(),
        })
    logger.info(f"Crawler | fetched {len(out)} agencies")
    return out


async def fetch_agency_xlsx(client: httpx.AsyncClient, agency_id: str) -> bytes | None:
    """GET export_exel_list_tthc.jsp?impl_agency_id=<id> → xlsx bytes."""
    try:
        r = await client.get(
            EXPORT_LIST_URL,
            params={
                "is_connected": 0,
                "keyword": "",
                "agency_type": 0,
                "impl_agency_id": agency_id,
                "object_id": -1,
                "field_id": -1,
                "impl_level_id": -1,
            },
            headers={**_DEFAULT_HEADERS, "Accept": "*/*"},
            timeout=settings.CRAWLER_TIMEOUT * 2,  # file lớn hơn, cho timeout dài hơn
        )
        r.raise_for_status()
        if not r.content.startswith(b"PK"):
            logger.warning(f"Crawler | non-xlsx agency list | agency_id={agency_id}")
            return None
        return r.content
    except Exception as e:
        logger.warning(f"Crawler | fetch_agency_xlsx failed | agency_id={agency_id} | {e}")
        return None


async def collect_all_codes_online(
    client: httpx.AsyncClient,
    agency_id: str | None = None,
) -> list[dict]:
    """
    [ONLINE] Lấy danh sách mã TTHC qua API (không cần file local).

    - agency_id=None → lấy TẤT CẢ cơ quan
    - agency_id="6369" → chỉ cơ quan đó
    """
    if agency_id:
        agencies = [{"id": agency_id, "name": "", "code": ""}]
    else:
        agencies = await fetch_agency_list(client)

    seen: set[str] = set()
    out: list[dict] = []
    for ag in agencies:
        xlsx_bytes = await fetch_agency_xlsx(client, ag["id"])
        if not xlsx_bytes:
            continue
        try:
            rows = read_codes_from_bytes(xlsx_bytes)
        except Exception as e:
            logger.warning(f"Crawler | parse agency xlsx failed | agency={ag['name']} | {e}")
            continue
        added = 0
        for r in rows:
            if r["code"] in seen:
                continue
            seen.add(r["code"])
            r["source_xlsx"] = ag["name"] or ag["id"]
            r["agency_id"] = ag["id"]
            out.append(r)
            added += 1
        logger.info(f"Crawler | agency '{ag['name'] or ag['id']}' | +{added} codes")
    logger.info(f"Crawler | collected {len(out)} unique codes from {len(agencies)} agencies (online)")
    return out


async def resolve_agency_id(client: httpx.AsyncClient, name_or_id: str) -> str | None:
    """
    Cho phép source_url là tên cơ quan (vd 'Bộ Công an') hoặc ID số.
    Trả về agency_id, hoặc None nếu không khớp.
    """
    s = name_or_id.strip()
    if s.isdigit():
        return s
    agencies = await fetch_agency_list(client)
    s_low = s.lower()
    # exact match trước, rồi substring
    for a in agencies:
        if a["name"].lower() == s_low:
            return a["id"]
    for a in agencies:
        if s_low in a["name"].lower():
            return a["id"]
    return None


# ── HTTP fetch ────────────────────────────────────────────────────────────────

async def _lookup_id_tthc(client: httpx.AsyncClient, code: str) -> str | None:
    """POST rest.jsp with keyword=<code> → return ID field (idTTHC)."""
    payload = {
        "service": "procedure_advanced_search_service_v2",
        "provider": "dvcquocgia",
        "type": "ref",
        "recordPerPage": 10,
        "pageIndex": 1,
        "is_connected": 0,
        "keyword": code,
        "agency_type": "0",
        "impl_agency_id": "-1",
        "object_id": "-1",
        "field_id": "-1",
        "impl_level_id": "-1",
    }
    try:
        r = await client.post(
            REST_URL,
            data={"params": json.dumps(payload, ensure_ascii=False)},
            headers=_DEFAULT_HEADERS,
            timeout=settings.CRAWLER_TIMEOUT,
        )
        r.raise_for_status()
        body = r.json()
    except Exception as e:
        logger.warning(f"Crawler | rest.jsp failed | code={code} | {e}")
        return None

    if not isinstance(body, list) or not body:
        logger.warning(f"Crawler | empty rest.jsp response | code={code}")
        return None

    # Find exact code match (response may include partial matches if multiple)
    for rec in body:
        if str(rec.get("PROCEDURE_CODE", "")).strip() == code:
            return str(rec.get("ID", "")).strip() or None

    # Fallback: first row
    return str(body[0].get("ID", "")).strip() or None


async def _download_docx(client: httpx.AsyncClient, code: str, id_tthc: str) -> bytes | None:
    """GET export_word_detail_tthc.jsp → docx bytes."""
    try:
        r = await client.get(
            EXPORT_URL,
            params={"maTTHC": code, "idTTHC": id_tthc},
            headers={**_DEFAULT_HEADERS, "Accept": "*/*"},
            timeout=settings.CRAWLER_TIMEOUT,
        )
        r.raise_for_status()
        # Sanity check: must be Office Open XML (PK header)
        if not r.content.startswith(b"PK"):
            logger.warning(
                f"Crawler | non-docx response | code={code} | first8={r.content[:8]!r}"
            )
            return None
        return r.content
    except Exception as e:
        logger.warning(f"Crawler | docx download failed | code={code} | id={id_tthc} | {e}")
        return None


# ── Orchestrator ──────────────────────────────────────────────────────────────

async def fetch_form_urls(client: httpx.AsyncClient, code: str) -> dict[str, str]:
    """
    Tải trang chi tiết server-rendered → map {form_filename: download_url}.
    download_url dạng https://csdl.dichvucong.gov.vn/web/jsp/download_file.jsp?ma=<hash>
    Trả về dict rỗng nếu không lấy được (không chặn pipeline).
    """
    try:
        r = await client.get(
            DETAIL_HTML_URL,
            params={"ma_thu_tuc": code},
            headers={**_DEFAULT_HEADERS, "Accept": "text/html"},
            timeout=settings.CRAWLER_TIMEOUT,
        )
        r.raise_for_status()
    except Exception as e:
        logger.warning(f"Crawler | fetch detail HTML failed | code={code} | {e}")
        return {}

    out: dict[str, str] = {}
    for url, name in _FORM_LINK_RE.findall(r.text):
        clean_name = ihtml.unescape(name).strip()
        if clean_name and clean_name not in out:
            out[clean_name] = url
    return out


async def fetch_and_parse_procedure(
    client: httpx.AsyncClient,
    code: str,
    fallback_meta: dict | None = None,
) -> dict | None:
    """
    Fetch a single procedure end-to-end:
      code → idTTHC → docx → parsed dict (with xlsx metadata as fallback)
      + lấy link tải biểu mẫu từ trang chi tiết HTML → gắn vào requirements[].form_url
    """
    id_tthc = await _lookup_id_tthc(client, code)
    if not id_tthc:
        return None

    docx_bytes = await _download_docx(client, code, id_tthc)
    if not docx_bytes:
        return None

    try:
        parsed = parse_docx(docx_bytes)
    except Exception as e:
        logger.warning(f"Crawler | parse_docx failed | code={code} | {e}")
        return None

    # Gắn link tải biểu mẫu: match form_name (filename) với map từ trang chi tiết
    form_urls = await fetch_form_urls(client, code)
    if form_urls:
        matched = 0
        for req in parsed.get("requirements", []):
            fname = (req.get("form_name") or "").strip()
            if fname and fname in form_urls:
                req["form_url"] = form_urls[fname]
                matched += 1
        logger.info(f"Crawler | form links | code={code} | matched {matched}/{len(form_urls)}")

    # Inject metadata from xlsx as fallback when docx doesn't have it
    if fallback_meta:
        parsed.setdefault("source_xlsx", fallback_meta.get("source_xlsx"))
        if not parsed.get("name") and fallback_meta.get("name_xlsx"):
            parsed["name"] = fallback_meta["name_xlsx"]
        if not parsed.get("domain") and fallback_meta.get("domain_xlsx"):
            parsed["domain"] = fallback_meta["domain_xlsx"]
        if not parsed.get("implementing_agency") and fallback_meta.get("agency_xlsx"):
            parsed["implementing_agency"] = fallback_meta["agency_xlsx"]
        if not parsed.get("authority_level_text") and fallback_meta.get("level_xlsx"):
            parsed["authority_level_text"] = fallback_meta["level_xlsx"]
        if not parsed.get("object") and fallback_meta.get("object_xlsx"):
            parsed["object"] = fallback_meta["object_xlsx"]
        if not parsed.get("decision_number") and fallback_meta.get("decision_xlsx"):
            parsed["decision_number"] = fallback_meta["decision_xlsx"]

    parsed["id_tthc"] = id_tthc
    return parsed


async def fetch_procedures(
    codes: Iterable[dict],
    concurrency: int = 5,
) -> AsyncIterator[tuple[str, dict | None]]:
    """
    Async generator: yields (code, parsed_or_None) for each code in `codes`.
    Bounded concurrency via semaphore.
    """
    sem = asyncio.Semaphore(concurrency)
    codes_list = list(codes)

    async with httpx.AsyncClient(http2=False, follow_redirects=True) as client:

        async def _one(meta: dict) -> tuple[str, dict | None]:
            async with sem:
                code = meta["code"]
                parsed = await fetch_and_parse_procedure(client, code, fallback_meta=meta)
                return code, parsed

        # asyncio.as_completed preserves async streaming
        tasks = [asyncio.create_task(_one(m)) for m in codes_list]
        for fut in asyncio.as_completed(tasks):
            yield await fut
